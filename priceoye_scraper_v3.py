"""
PriceOye Installment Scraper v3 — PRODUCTION GRADE
═════════════════════════════════════════════════════
Ultra-fast, accurate, anti-bot bypass, JS extraction.

INSTALL (one-time):
    pip install undetected-chromedriver selenium openpyxl beautifulsoup4 lxml rich
    
USAGE:
    python priceoye_scraper_v3.py "https://priceoye.pk/mobiles/apple/apple-iphone-16"
"""

import sys, re, time, json, random, asyncio, argparse, logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List
from concurrent.futures import ThreadPoolExecutor
import warnings
warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.WARNING)

# ══════════════════════════════════════════════════════════════════════════════
# DEPENDENCY CHECK
# ══════════════════════════════════════════════════════════════════════════════
missing = []
for pkg, imp in [
    ("undetected-chromedriver", "undetected_chromedriver"),
    ("selenium", "selenium"),
    ("openpyxl", "openpyxl"),
    ("bs4", "bs4"),
]:
    try:
        __import__(imp)
    except ImportError:
        missing.append(pkg)

if missing:
    print(f"❌ Missing packages: pip install {' '.join(missing)}")
    sys.exit(1)

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

try:
    from rich.console import Console
    from rich.table import Table as RichTable
    from rich.progress import Progress, SpinnerColumn, TextColumn
    C = Console()
    HAS_RICH = True
except ImportError:
    HAS_RICH = False

def info(m):  print(f"  ➤ {m}") if not HAS_RICH else C.print(f"[cyan]➤[/cyan] {m}")
def ok(m):    print(f"  ✓ {m}") if not HAS_RICH else C.print(f"[green]✓[/green] {m}")
def warn(m):  print(f"  ⚠ {m}") if not HAS_RICH else C.print(f"[yellow]⚠[/yellow] {m}")
def err(m):   print(f"  ✗ {m}") if not HAS_RICH else C.print(f"[red]✗[/red] {m}")
def head(m):  print(f"\n{'='*60}\n{m}\n{'='*60}") if not HAS_RICH else C.print(f"\n[bold cyan]{m}[/bold cyan]")

BASE = "https://priceoye.pk"

# All known banks on PriceOye with their BNPL pages
BANKS = {
    "HBL": f"{BASE}/bnpl-hbl",
    "MCB": f"{BASE}/bnpl-mcb",
    "UBL": f"{BASE}/bnpl-ubl",
    "Bank Alfalah": f"{BASE}/bnpl-alfalah",
    "Faysal Bank": f"{BASE}/bnpl-faysal",
    "Askari Bank": f"{BASE}/bnpl-askari",
    "Silk Bank": f"{BASE}/bnpl-silk",
    "JS Bank": f"{BASE}/bnpl-js",
    "Meezan Bank": f"{BASE}/bnpl-meezan",
    "Standard Chartered": f"{BASE}/bnpl-sc",
    "Habib Metro": f"{BASE}/bnpl-habibmetro",
}

BANK_COLORS = {
    "HBL":"1B4F72", "MCB":"154360", "UBL":"186A3B",
    "Bank Alfalah":"7D6608", "Faysal Bank":"6E2F1A",
    "Askari Bank":"512E5F", "Silk Bank":"1B2631",
    "JS Bank":"0E6655", "Meezan Bank":"784212",
    "Standard Chartered":"1F618D", "Habib Metro":"17202A",
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

# ══════════════════════════════════════════════════════════════════════════════
# BROWSER MANAGER — undetected-chromedriver
# ══════════════════════════════════════════════════════════════════════════════
class ChromeBrowser:
    def __init__(self):
        self.driver = None

    def start(self, headless=True):
        """Launch undetected Chrome with anti-bot measures."""
        options = uc.ChromeOptions()
        options.user_agent = random.choice(USER_AGENTS)
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1366,768")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-plugins")
        prefs = {"profile.managed_default_content_settings.images": 2}
        options.add_experimental_option("prefs", prefs) # disable images for speed
        
        if headless:
            options.add_argument("--headless=new")

        self.driver = uc.Chrome(options=options, version_main=None, suppress_welcome=True)
        self.driver.set_page_load_timeout(45)
        self.driver.implicitly_wait(10)
        ok("Browser launched (undetected-chromedriver)")
        return self

    def get(self, url: str, wait_selector: str = None, wait_ms: int = 4000):
        """Load a URL and optionally wait for an element."""
        try:
            info(f"Loading: {url}")
            self.driver.get(url)

            if wait_selector:
                try:
                    WebDriverWait(self.driver, 15).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, wait_selector))
                    )
                except TimeoutException:
                    warn(f"Selector not found: {wait_selector}, continuing anyway")

            # let JS render
            time.sleep(wait_ms / 1000)

            # scroll to trigger lazy loading
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            self.driver.execute_script("window.scrollTo(0, 0);")

            return self.driver.page_source

        except Exception as e:
            err(f"Failed to load {url}: {e}")
            return ""

    def execute_js(self, script: str):
        """Execute JavaScript and return result."""
        try:
            return self.driver.execute_script(script)
        except Exception as e:
            warn(f"JS execution failed: {e}")
            return None

    def close(self):
        """Close the browser."""
        try:
            if self.driver:
                self.driver.quit()
                ok("Browser closed")
        except Exception as e:
            warn(f"Error closing browser: {e}")

    def __enter__(self):
        return self.start()

    def __exit__(self, *_):
        self.close()


# ══════════════════════════════════════════════════════════════════════════════
# SCRAPER — Product page + Bank pages
# ══════════════════════════════════════════════════════════════════════════════
def scrape_product(driver, product_url: str) -> tuple[str, float]:
    """Extract product name and price from product page."""
    html = driver.get(product_url, wait_selector="h1", wait_ms=5000)
    soup = BeautifulSoup(html, "lxml")

    # Extract name
    name = ""
    if soup.find("h1"):
        name = soup.find("h1").get_text(strip=True)
    if not name:
        name = product_url.split("/")[-1].replace("-", " ").title()

    # Extract price using multiple strategies
    price = 0.0

    # Strategy 1: look for price in common patterns
    for sel in ['[class*="price"]', '[data-price]', '.price']:
        for el in soup.select(sel):
            text = el.get_text(" ", strip=True)
            # match "309,999" or "309999"
            matches = re.findall(r'[\d,]+', text)
            for m in matches:
                try:
                    val = float(m.replace(",", ""))
                    if 50000 < val < 10_000_000:
                        price = val
                        break
                except ValueError:
                    pass
            if price:
                break
        if price:
            break

    # Strategy 2: regex over full text
    if not price:
        text = soup.get_text()
        # "Rs. 309,999" or "PKR 309,999"
        for m in re.finditer(r'(?:Rs\.?|PKR)\s*([\d,]+)', text):
            try:
                val = float(m.group(1).replace(",", ""))
                if 50000 < val < 10_000_000:
                    price = val
                    break
            except ValueError:
                pass

    ok(f"Product: {name} | Price: PKR {price:,.0f}" if price else f"Product: {name} | Price: NOT FOUND")
    return name, price


JS_EXTRACT_INSTALLMENTS = """
// Extract installment table from the page using JS (faster than HTML parsing)
let results = [];

// Strategy 1: Look for table with installment data
let tables = document.querySelectorAll('table');
for (let table of tables) {
    let text = table.innerText.toLowerCase();
    if (!text.includes('month') && !text.includes('installment') && !text.includes('tenure')) {
        continue;
    }
    
    let rows = table.querySelectorAll('tr');
    if (rows.length < 2) continue;
    
    // Get headers
    let headers = [];
    rows[0].querySelectorAll('th, td').forEach(cell => {
        headers.push(cell.innerText.toLowerCase().trim());
    });
    
    // Parse data rows
    for (let i = 1; i < rows.length; i++) {
        let cells = [];
        rows[i].querySelectorAll('td, th').forEach(cell => {
            cells.push(cell.innerText.trim());
        });
        
        if (cells.length === 0 || cells.every(c => c === '')) continue;
        
        // Find tenure and amount columns
        let tenure = null;
        let amount = null;
        
        for (let j = 0; j < cells.length; j++) {
            let cell = cells[j];
            let header = headers[j] || '';
            
            // Tenure column
            if (header.includes('month') || header.includes('tenure') || header.includes('installment plan')) {
                let match = cell.match(/\\d+/);
                if (match) {
                    tenure = parseInt(match[0]);
                }
            }
            
            // Amount column (look for Rs, PKR, or large numbers)
            if (header.includes('amount') || header.includes('installment') || header.includes('emi') || 
                header.includes('price') || header.includes('monthly')) {
                let match = cell.match(/[\\d,]+/);
                if (match) {
                    amount = parseInt(match[0].replace(/,/g, ''));
                }
            }
        }
        
        if (tenure && amount) {
            results.push({ tenure: tenure, amount: amount });
        }
    }
    
    if (results.length > 0) {
        return results;
    }
}

// Strategy 2: Look for div-based layouts with data attributes
let sections = document.querySelectorAll('[class*="installment"], [class*="emi"], [class*="payment"]');
for (let section of sections) {
    let text = section.innerText.toLowerCase();
    if (!text.match(/\\d+.*month|tenure/i)) continue;
    
    // Look for text patterns like "3 Months Rs 107,494"
    let matches = section.innerText.matchAll(/([\\d]+)\\s*months?[^\\d]*([\\.\\d,]+)/gi);
    for (let match of matches) {
        let tenure = parseInt(match[1]);
        let amountStr = match[2].replace(/[^\\d.]/g, '');
        let amount = parseInt(amountStr.replace(/\\./g, '').substring(0, 6));
        if (tenure && amount > 1000) {
            results.push({ tenure: tenure, amount: amount });
        }
    }
}

return results.length > 0 ? results : [];
"""


def scrape_bank_rates(driver, bank_name: str, bnpl_url: str) -> List[Dict]:
    """Scrape installment data from bank BNPL page using JS extraction."""
    info(f"Scraping {bank_name}...")
    html = driver.get(bnpl_url, wait_selector="table, [class*='installment']", wait_ms=4500)

    if not html or len(html) < 1000:
        warn(f"  {bank_name}: Empty page, skipping")
        return []

    # Try JS extraction first (FASTEST)
    results = driver.execute_js(JS_EXTRACT_INSTALLMENTS)

    if results and len(results) > 0:
        # Deduplicate
        seen = set()
        unique = []
        for r in results:
            key = (r['tenure'], r['amount'])
            if key not in seen:
                seen.add(key)
                unique.append(r)
        ok(f"  {bank_name}: {len(unique)} plans extracted")
        return unique

    # Fallback: HTML parsing
    warn(f"  {bank_name}: JS extraction failed, trying HTML parse")
    soup = BeautifulSoup(html, "lxml")
    results = _parse_html_table(soup)

    if results:
        ok(f"  {bank_name}: {len(results)} plans found (HTML parse)")
        return results

    warn(f"  {bank_name}: No data found")
    return []


def _parse_html_table(soup) -> List[Dict]:
    """Fallback HTML table parser."""
    results = []
    for table in soup.find_all("table"):
        text = table.get_text().lower()
        if not any(k in text for k in ["month", "installment", "emi", "tenure"]):
            continue

        rows = table.find_all("tr")
        if len(rows) < 2:
            continue

        for row in rows[1:]:
            cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
            if len(cells) < 2:
                continue

            # Try to extract tenure and amount
            tenure = None
            amount = None

            for cell in cells:
                m_tenure = re.search(r'\b([3-9]|[1-2]\d)\s*(?:months?|month)\b', cell, re.I)
                if m_tenure:
                    tenure = int(m_tenure.group(1))

                m_amount = re.search(r'([\d,]+)', cell)
                if m_amount:
                    try:
                        val = int(m_amount.group(1).replace(",", ""))
                        if 10000 < val < 500000:
                            amount = val
                    except ValueError:
                        pass

            if tenure and amount:
                results.append({"tenure": tenure, "amount": amount})

    # Deduplicate
    seen = set()
    unique = []
    for r in results:
        key = (r['tenure'], r['amount'])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    return unique


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(records: List[Dict], product_name: str, price: float, path: str = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "All Plans"

    # Styles
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill("solid", start_color="1A3F5C")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DAT_ALIGN = Alignment(horizontal="center", vertical="center")
    ALT_FILL = PatternFill("solid", start_color="D6EAF8")
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1A3F5C")

    # Title
    N = 9
    LC = get_column_letter(N)
    ws.merge_cells(f"A1:{LC}1")
    ws["A1"] = f"PriceOye Installment Plans — {product_name}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{LC}2")
    ws["A2"] = f"Device Price: PKR {price:,.0f}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  All amounts in PKR"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Header
    COLS = ["Bank", "Tenure (Months)", "Monthly Installment (PKR)", "Total Payable (PKR)", "Down Payment (PKR)"]
    HR = 4
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(HR, ci, col)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border = BORDER
    ws.row_dimensions[HR].height = 22

    # Data (grouped by bank)
    ri = HR + 1
    cur_bank = None
    alt = False

    for rec in records:
        bank = rec["bank"]
        tenure = rec["tenure"]
        amount = rec["amount"]
        total = tenure * amount

        # Bank separator
        if bank != cur_bank:
            cur_bank = bank
            alt = False
            ws.merge_cells(f"A{ri}:{LC}{ri}")
            bfill = PatternFill("solid", start_color=BANK_COLORS.get(bank, "2C3E50"))
            c = ws.cell(ri, 1, f"  {bank}")
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            c.fill = bfill
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[ri].height = 18
            ri += 1

        # Data cells
        fill = ALT_FILL if alt else None
        data = [bank, tenure, amount, total, 0]  # down payment = 0 for BNPL
        for ci, val in enumerate(data, 1):
            c = ws.cell(ri, ci, val)
            c.border = BORDER
            c.alignment = DAT_ALIGN
            if fill:
                c.fill = fill
            if isinstance(val, int) and ci > 2:
                c.number_format = '#,##0'

        ws.row_dimensions[ri].height = 16
        alt = not alt
        ri += 1

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 20

    ws.freeze_panes = f"A{HR + 1}"
    ws.auto_filter.ref = f"A{HR}:{LC}{ri-1}"

    # Summary sheet
    ws2 = wb.create_sheet("Best Deals")
    ws2["A1"] = f"Best Deals — Lowest Monthly Installment"
    ws2["A1"].font = TITLE_FONT
    sorted_recs = sorted(records, key=lambda r: r["amount"])
    for ri, rec in enumerate(sorted_recs, 3):
        ws2[f"A{ri}"] = rec["bank"]
        ws2[f"B{ri}"] = rec["tenure"]
        ws2[f"C{ri}"] = rec["amount"]
        for col in ["A", "B", "C"]:
            ws2[f"{col}{ri}"].border = BORDER
            ws2[f"{col}{ri}"].alignment = DAT_ALIGN
            if col in ["B", "C"] and isinstance(rec.get(col.lower(), 0), int):
                ws2[f"{col}{ri}"].number_format = '#,##0'

    if not path:
        safe = re.sub(r'[\\/*?:"<>|]', "_", product_name)[:50]
        path = f"priceoye_{safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main(product_url: str, output_path: str = None):
    head("PriceOye Installment Scraper v3 — PRODUCTION")

    with ChromeBrowser() as driver:
        # Step 1: Product page
        product_name, price = scrape_product(driver, product_url)

        if not price:
            try:
                price = float(input("\n  Price not detected. Enter manually (PKR): ").strip().replace(",", ""))
            except (ValueError, KeyboardInterrupt):
                err("No price. Exiting.")
                sys.exit(1)

        # Step 2: Scrape all bank BNPL pages
        head("Scraping bank installment plans...")
        all_records = []

        for bank_name, bnpl_url in BANKS.items():
            plans = scrape_bank_rates(driver, bank_name, bnpl_url)

            for plan in plans:
                all_records.append({
                    "bank": bank_name,
                    "tenure": plan["tenure"],
                    "amount": plan["amount"],
                })

            time.sleep(random.uniform(0.5, 1.5))  # polite delay

    if not all_records:
        err("No installment data found. Check your internet.")
        sys.exit(1)

    ok(f"\nTotal: {len(all_records)} plans from {len(set(r['bank'] for r in all_records))} banks")

    # Step 3: Export
    head("Exporting to Excel...")
    xlsx = export_excel(all_records, product_name, price, output_path)
    ok(f"✓ Saved: {xlsx}")

    return xlsx


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PriceOye Installment Scraper v3")
    parser.add_argument("url", nargs="?", help="PriceOye product URL")
    parser.add_argument("-o", "--output", help="Output Excel file")
    args = parser.parse_args()

    url = args.url
    if not url:
        print("\nPriceOye Installment Scraper v3")
        url = input("Product URL: ").strip()
        if not url:
            print("No URL provided.")
            sys.exit(1)

    try:
        main(url, args.output)
    except KeyboardInterrupt:
        print("\nCancelled.")
    except Exception as e:
        err(str(e))
        import traceback
        traceback.print_exc()
        sys.exit(1)